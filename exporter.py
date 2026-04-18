"""
exporter.py - Multi-Format Output Export
Support for JSON, CSV, Excel with advanced formatting
"""

import json
import csv
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime


class BaseExporter:
    """Base exporter class with common functionality."""
    
    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def _get_filename(self, base_name: str, extension: str) -> str:
        """Generate filename with timestamp."""
        from config import get_config
        config = get_config()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if config.get('output.timestamp_output') else ""
        if timestamp:
            return f"{base_name}_{timestamp}.{extension}"
        return f"{base_name}.{extension}"


class JSONExporter(BaseExporter):
    """Export data to JSON format."""
    
    def export(self, data: Dict, filename: str) -> str:
        """Export data to JSON file."""
        from config import get_config
        config = get_config()
        
        output_file = self.output_dir / self._get_filename(filename, 'json')
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(
                data,
                f,
                indent=config.get('output.indent', 2),
                ensure_ascii=config.get('output.ensure_ascii', False)
            )
        
        return str(output_file)


class CSVExporter(BaseExporter):
    """Export data to CSV format."""
    
    def export(self, data: Dict, filename: str) -> str:
        """Export summary data to CSV file."""
        output_file = self.output_dir / self._get_filename(filename, 'csv')
        
        try:
            rows = data.get('summary', {}).get('rows', [])
            if not rows:
                raise ValueError("No data to export")
            
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                fieldnames = ['pos', 'taxable_value', 'igst', 'cgst', 'sgst']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for row in rows:
                    filtered_row = {k: row.get(k, '') for k in fieldnames}
                    writer.writerow(filtered_row)
            
            return str(output_file)
        except Exception as e:
            raise ValueError(f"CSV export failed: {str(e)}")


class ExcelExporter(BaseExporter):
    """Export data to Excel format with formatting."""
    
    def export(self, data: Dict, filename: str) -> str:
        """Export data to Excel file with formatting."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")
        
        output_file = self.output_dir / self._get_filename(filename, 'xlsx')
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        
        summary = data.get('summary', {})
        rows = summary.get('rows', [])
        
        # Headers
        headers = ['POS', 'Taxable Value', 'IGST', 'CGST', 'SGST']
        ws_summary.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Data rows
        for row in rows:
            ws_summary.append([
                row.get('pos', ''),
                row.get('taxable_value', 0),
                row.get('igst', 0),
                row.get('cgst', 0),
                row.get('sgst', 0)
            ])
        
        # Totals
        summary_total = data.get('summary', {})
        ws_summary.append([
            'TOTAL',
            summary_total.get('total_taxable', 0),
            summary_total.get('total_igst', 0),
            summary_total.get('total_cgst', 0),
            summary_total.get('total_sgst', 0)
        ])
        
        # Auto-adjust column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_summary.column_dimensions[column_letter].width = max_length + 2
        
        # Credit docs sheet
        if data.get('credit_docs'):
            ws_credit = wb.create_sheet('Credit Docs')
            ws_credit.append(['Invoice No', 'POS', 'Taxable Value', 'IGST', 'CGST', 'SGST'])
            for doc in data['credit_docs']:
                ws_credit.append([
                    doc.get('invoice_no', ''),
                    doc.get('pos', ''),
                    doc.get('txval', 0),
                    doc.get('igst_amt', 0),
                    doc.get('cgst_amt', 0),
                    doc.get('sgst_amt', 0)
                ])
        
        wb.save(output_file)
        return str(output_file)


class Exporter:
    """Multi-format exporter orchestrator."""
    
    def __init__(self, output_dir: str = "./output"):
        self.json_exporter = JSONExporter(output_dir)
        self.csv_exporter = CSVExporter(output_dir)
        self.excel_exporter = ExcelExporter(output_dir)
    
    def export(self, data: Dict, base_filename: str, formats: List[str] = None) -> Dict[str, str]:
        """
        Export data to multiple formats.
        
        Args:
            data: Data to export
            base_filename: Base filename without extension
            formats: List of formats ('json', 'csv', 'xlsx')
        
        Returns:
            Dictionary mapping format to file path
        """
        from config import get_config
        config = get_config()
        
        if formats is None:
            formats = config.get('output.formats', ['json'])
        
        results = {}
        for fmt in formats:
            try:
                if fmt == 'json':
                    results['json'] = self.json_exporter.export(data, base_filename)
                elif fmt == 'csv':
                    results['csv'] = self.csv_exporter.export(data, base_filename)
                elif fmt == 'xlsx' or fmt == 'excel':
                    results['xlsx'] = self.excel_exporter.export(data, base_filename)
            except Exception as e:
                results[fmt] = f"ERROR: {str(e)}"
        
        return results
